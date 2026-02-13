"""
Document extractor module for AgentCore Execution Agent.

Extracts text content from various document formats:
- PDF: PyPDF2
- DOCX: XML parsing (standard library, no lxml dependency)
- CSV: built-in csv module
- XLSX: openpyxl
- PPTX: XML parsing (standard library, no lxml dependency)
- TXT: built-in file reading

Adapted from Lambda version for AgentCore container environment.

Note: DOCX and PPTX use XML parsing directly (no lxml dependency).
PPTX slide-to-image conversion is not supported (LibreOffice removed).
"""

import csv
import json
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO, StringIO
from typing import List, Optional

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


from logger_util import get_logger, log

_logger = get_logger()


def _log(level: str, event_type: str, data: dict) -> None:
    """Structured JSON logging for CloudWatch."""
    log(_logger, level, event_type, data, service="execution-agent-document-extractor")


def extract_text_from_pdf(pdf_bytes: bytes) -> Optional[str]:
    """
    Extract text from PDF file.

    Args:
        pdf_bytes: PDF file content as bytes

    Returns:
        Extracted text as string, or None if extraction fails
    """
    if PyPDF2 is None:
        return None

    try:
        pdf_file = BytesIO(pdf_bytes)
        pdf_reader = PyPDF2.PdfReader(pdf_file)

        text_parts = []
        for page in pdf_reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)

        return "\n\n".join(text_parts) if text_parts else None
    except Exception as e:
        _log("ERROR", "pdf_extraction_failed", {"error": str(e), "error_type": type(e).__name__})
        return None


def _extract_text_from_docx_xml(docx_bytes: bytes) -> Optional[str]:
    """
    Extract text from DOCX file by parsing XML directly (no lxml dependency).

    DOCX files are ZIP archives containing XML files. This function extracts
    text from the main document XML (word/document.xml).

    Args:
        docx_bytes: DOCX file content as bytes

    Returns:
        Extracted text as string, or None if extraction fails
    """
    try:
        docx_zip = zipfile.ZipFile(BytesIO(docx_bytes))

        # DOCX namespace
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

        text_parts = []

        # Extract text from main document
        if "word/document.xml" in docx_zip.namelist():
            doc_xml = docx_zip.read("word/document.xml")
            root = ET.fromstring(doc_xml)

            # Find all text elements (w:t)
            for text_elem in root.findall(".//w:t", ns):
                if text_elem.text:
                    text_parts.append(text_elem.text)

        return "\n".join(text_parts) if text_parts else None
    except Exception as e:
        _log("WARN", "docx_xml_extraction_failed", {"error": str(e), "error_type": type(e).__name__})
        return None


def extract_text_from_docx(docx_bytes: bytes) -> Optional[str]:
    """
    Extract text from DOCX file using XML parsing.

    Args:
        docx_bytes: DOCX file content as bytes

    Returns:
        Extracted text as string, or None if extraction fails
    """
    return _extract_text_from_docx_xml(docx_bytes)


def extract_text_from_csv(csv_bytes: bytes) -> Optional[str]:
    """
    Extract text from CSV file.

    Args:
        csv_bytes: CSV file content as bytes

    Returns:
        Extracted text as string, or None if extraction fails
    """
    try:
        csv_string = csv_bytes.decode("utf-8", errors="replace")
        csv_file = StringIO(csv_string)
        reader = csv.reader(csv_file)

        rows = []
        for row in reader:
            rows.append(",".join(str(cell) for cell in row))

        return "\n".join(rows) if rows else None
    except Exception as e:
        _log("ERROR", "csv_extraction_failed", {"error": str(e), "error_type": type(e).__name__})
        return None


def extract_text_from_xlsx(xlsx_bytes: bytes) -> Optional[str]:
    """
    Extract text from XLSX file.

    Args:
        xlsx_bytes: XLSX file content as bytes

    Returns:
        Extracted text as string, or None if extraction fails
    """
    if openpyxl is None:
        return None

    try:
        xlsx_file = BytesIO(xlsx_bytes)
        workbook = openpyxl.load_workbook(xlsx_file, data_only=True)

        text_parts = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}")

            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                if row_text.strip():
                    text_parts.append(row_text)

            text_parts.append("")  # Separator between sheets

        return "\n".join(text_parts) if text_parts else None
    except Exception as e:
        _log("ERROR", "xlsx_extraction_failed", {"error": str(e), "error_type": type(e).__name__})
        return None


def _extract_text_from_pptx_xml(pptx_bytes: bytes) -> Optional[str]:
    """
    Extract text from PPTX file by parsing XML directly (no lxml dependency).

    PPTX files are ZIP archives containing XML files. This function extracts
    text from slide XML files (ppt/slides/slide*.xml).

    Args:
        pptx_bytes: PPTX file content as bytes

    Returns:
        Extracted text as string, or None if extraction fails
    """
    try:
        pptx_zip = zipfile.ZipFile(BytesIO(pptx_bytes))

        # PPTX namespace
        ns = {
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
        }

        text_parts = []
        slide_num = 0

        # Find all slide XML files
        slide_files = [
            f
            for f in pptx_zip.namelist()
            if f.startswith("ppt/slides/slide") and f.endswith(".xml")
        ]
        slide_files.sort()

        for slide_file in slide_files:
            slide_num += 1
            slide_xml = pptx_zip.read(slide_file)
            root = ET.fromstring(slide_xml)

            slide_texts = []

            # Find all text elements (a:t)
            for text_elem in root.findall(".//a:t", ns):
                if text_elem.text:
                    slide_texts.append(text_elem.text)

            if slide_texts:
                text_parts.append(f"Slide {slide_num}:")
                text_parts.extend(slide_texts)
                text_parts.append("")

        return "\n".join(text_parts) if text_parts else None
    except Exception as e:
        _log("WARN", "pptx_xml_extraction_failed", {"error": str(e), "error_type": type(e).__name__})
        return None


def extract_text_from_pptx(pptx_bytes: bytes) -> Optional[str]:
    """
    Extract text from PPTX file using XML parsing.

    Args:
        pptx_bytes: PPTX file content as bytes

    Returns:
        Extracted text as string, or None if extraction fails
    """
    return _extract_text_from_pptx_xml(pptx_bytes)


def convert_pptx_slides_to_images(pptx_bytes: bytes) -> Optional[List[bytes]]:
    """
    Convert PPTX slides to PNG images.

    NOTE: This function is currently disabled. LibreOffice-based conversion
    has been removed. PPTX files will only have text extraction.

    Args:
        pptx_bytes: PPTX file content as bytes

    Returns:
        Always returns None (feature disabled)
    """
    return None


def extract_text_from_txt(txt_bytes: bytes) -> Optional[str]:
    """
    Extract text from TXT file.

    Args:
        txt_bytes: TXT file content as bytes

    Returns:
        Extracted text as string, or None if extraction fails
    """
    try:
        return txt_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        _log("ERROR", "txt_extraction_failed", {"error": str(e), "error_type": type(e).__name__})
        return None
