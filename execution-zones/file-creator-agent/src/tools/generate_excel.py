"""
generate_excel tool for Execution Agent (027-slack-file-generation-best-practices).

Produces Excel (.xlsx) files and stores them in tool_context.invocation_state
for the handler to extract and build file_artifact.
"""

from io import BytesIO
from typing import Any, List

from strands import tool

import file_config as fc
from file_config import sanitize_filename

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@tool(context=True)
def generate_excel(
    filename: str,
    sheets: List[dict],
    tool_context: Any,
) -> str:
    """Excelスプレッドシート (.xlsx) を生成します。

    複数シート対応。各シートにヘッダーとデータ行を指定できます。

    Args:
        filename: ファイル名（拡張子不要、自動付加）
        sheets: シートの配列。各要素は name, headers, rows を持つ。
        tool_context: Strands framework context (injected). Contains invocation_state.

    Returns:
        モデルに返す説明文（日本語）。ファイルは invocation_state に格納される。
    """
    if not filename or not isinstance(filename, str) or not filename.strip():
        return "エラー: filename を指定してください。"
    if not sheets or not isinstance(sheets, list):
        return "エラー: sheets は必須です（name, headers, rows を持つオブジェクトの配列）。"

    if len(sheets) > 100:
        return "エラー: シート数は最大100までです。"

    safe_filename = sanitize_filename(filename.strip(), "xlsx")
    if not safe_filename or "." not in safe_filename:
        safe_filename = f"{safe_filename}.xlsx"

    try:
        from openpyxl import Workbook
    except ImportError:
        return "エラー: openpyxl がインストールされていません。"

    wb = Workbook()
    default_sheet = wb.active
    sheet_count = 0

    for sheet_idx, sheet_def in enumerate(sheets):
        if not isinstance(sheet_def, dict):
            continue
        name = sheet_def.get("name") or f"Sheet{sheet_idx + 1}"
        headers = sheet_def.get("headers")
        rows = sheet_def.get("rows")

        if not isinstance(headers, list):
            headers = []
        if not isinstance(rows, list):
            rows = []

        if sheet_count == 0:
            ws = default_sheet
            ws.title = str(name)[:31]
        else:
            ws = wb.create_sheet(title=str(name)[:31])
        sheet_count += 1

        for col, header in enumerate(headers, start=1):
            if header is not None:
                ws.cell(row=1, column=col, value=str(header))

        for row_idx, row in enumerate(rows, start=2):
            if not isinstance(row, list):
                continue
            for col_idx, cell_val in enumerate(row, start=1):
                if cell_val is not None:
                    ws.cell(row=row_idx, column=col_idx, value=cell_val)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.getvalue()

    if len(file_bytes) > fc.MAX_OFFICE_FILE_BYTES:
        return (
            f"エラー: ファイルサイズが上限（{fc.MAX_OFFICE_FILE_BYTES} バイト）を超えています。"
            f"現在 {len(file_bytes)} バイトです。"
        )

    invocation_state = getattr(tool_context, "invocation_state", {})
    if isinstance(invocation_state, dict):
        invocation_state["generated_file"] = {
            "file_bytes": file_bytes,
            "file_name": safe_filename,
            "mime_type": _XLSX_MIME,
            "description": f"Excelファイル「{safe_filename}」を生成しました。",
        }

    return f"ファイル「{safe_filename}」を作成しました。Slack にアップロードされます。"
